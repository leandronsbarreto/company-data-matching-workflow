from pathlib import Path
from datetime import datetime
import shutil
import re

from openpyxl import load_workbook
from openpyxl.styles import Alignment


# ============================================================
# SETTINGS
# ============================================================

BASE_DIR = Path(__file__).resolve().parent
SAMPLE_DATA_DIR = BASE_DIR / "sample_data"

MATCH_FILE = SAMPLE_DATA_DIR / "Match_Data.xlsx"
EXTRACTED_BATCH_FILE = SAMPLE_DATA_DIR / "Extracted_Batch.xlsx"

SEARCH_EXPORT_PATTERN = "Search_Export_*.xlsx"

SHEET_DATA = "Source_Data"
SHEET_NOT_FOUND = "Not_Found"
SHEET_LOG = "Apply_Log"

EURO_FORMAT = '#,##0.00 €'

# Search_Export columns
EXPORT_COL_INTERNAL_ID = 1          # A - Interne ID
EXPORT_COL_NAME = 2                 # B - Unternehmensname
EXPORT_COL_NATIONAL_ID = 3          # C - Nationale ID
EXPORT_COL_CITY = 4                 # D - Ort, not used for matching
EXPORT_COL_DATABASE_ID = 7          # G - Gematchte Datenbank-ID
EXPORT_COL_MATCHED_NAME = 8         # H - Gematchter Name, used only for previous-name extraction

# Match_Data columns, sheet Source_Data
TARGET_COL_INTERNAL_ID = 1          # A - Interne ID
TARGET_COL_NAME = 2                 # B - Unternehmensname
TARGET_COL_CITY = 3                 # C - Ort
TARGET_COL_NATIONAL_ID = 4          # D - Nationale ID
TARGET_PASTE_START_COL = 5          # E
TARGET_PASTE_END_COL = 21           # U
TARGET_COL_FORMER_NAME = 22         # V - Ehemaliger Name
TARGET_COL_DUP_GROUP = 23           # W - Duplikatgruppen-ID

# Extracted_Batch columns
BATCH_COL_NAME = 3                  # C - Unternehmensname
BATCH_COL_DATABASE_ID = 4           # D - Datenbank-ID
BATCH_COL_UID = 15                  # O - UID-Nummer


# ============================================================
# CLEANING AND MATCHING FUNCTIONS
# ============================================================

def clean_outer_quotes(value):
    """
    Remove leading and trailing quotation marks and surrounding spaces.

    Important for cases such as:
    "1001
    DE101000001"
    """
    if value is None:
        return ""

    text = str(value).strip()

    while text.startswith('"'):
        text = text[1:].strip()

    while text.endswith('"'):
        text = text[:-1].strip()

    return text


def remove_all_whitespace(value):
    """
    Remove all whitespace characters from a value.

    Examples:
    DE 101 000 001 -> DE101000001
    12 / 67 / 19 -> 12/67/19

    Slashes and other non-whitespace characters are preserved.
    """
    if value is None:
        return ""

    text = clean_outer_quotes(value)
    text = re.sub(r"\s+", "", text)
    return text


def collapse_whitespace(value):
    """
    Replace multiple spaces, tabs and line breaks with a single space.
    """
    if value is None:
        return ""

    text = clean_outer_quotes(value)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def clean_id_or_database_id_for_match(value):
    """
    Clean internal IDs and Datenbank-IDs for strict comparison.

    Removes only:
    - leading/trailing quotation marks;
    - surrounding spaces.

    It does not change internal characters.
    """
    return clean_outer_quotes(value)


def clean_national_id_for_match(value):
    """
    Clean Nationale ID / UID-Nummer values for matching.

    Removes:
    - leading/trailing quotation marks;
    - all whitespace characters.

    Used for:
    - Search_Export column C;
    - Source_Data column D;
    - Extracted_Batch column O.
    """
    return remove_all_whitespace(value)


def normalize_company_name_for_match(value):
    """
    Normalize company names for matching.

    This intentionally treats the following as equivalent:
    COMPANY RGTZP23 MBH
    COMPANY   RGTZ-P23 MBH
    Company RgtZp23 mbH

    Rules:
    - remove leading/trailing quotation marks;
    - ignore upper/lower case;
    - remove whitespace;
    - remove special characters;
    - keep letters and numbers.

    This is used only for matching. It does not overwrite the original
    company name in column B.
    """
    if value is None:
        return ""

    text = clean_outer_quotes(value)
    text = text.casefold()
    text = re.sub(r"[^0-9a-zäöüß]", "", text, flags=re.IGNORECASE)
    return text


def make_match_key(internal_id, name, national_id):
    """
    Build the secure matching key.

    Search_Export:
    A Interne ID + B Unternehmensname + C Nationale ID

    against Source_Data:
    A Interne ID + B Unternehmensname + D Nationale ID

    Interne ID is strict.
    Company name is normalized for spaces, case and special characters.
    Nationale ID is compared after removing all whitespace.
    """
    return (
        clean_id_or_database_id_for_match(internal_id),
        normalize_company_name_for_match(name),
        clean_national_id_for_match(national_id),
    )


# ============================================================
# FILE AND WORKSHEET FUNCTIONS
# ============================================================

def find_search_export_file():
    """
    Find exactly one Search_Export_[number].xlsx file in sample_data.

    The script stops if none or more than one is found.
    """
    files = sorted(SAMPLE_DATA_DIR.glob(SEARCH_EXPORT_PATTERN))

    files = [
        file_path for file_path in files
        if not file_path.name.startswith("~$")
        and file_path.is_file()
    ]

    if len(files) == 0:
        raise FileNotFoundError(
            f"No file found with the pattern {SEARCH_EXPORT_PATTERN} in {SAMPLE_DATA_DIR}."
        )

    if len(files) > 1:
        names = "\n".join(f"- {file_path.name}" for file_path in files)
        raise RuntimeError(
            "More than one search export file was found. "
            "Please keep only one Search_Export_[number].xlsx file in sample_data.\n\n"
            f"Files found:\n{names}"
        )

    return files[0]


def create_backup(file_path):
    """
    Create a timestamped backup of Match_Data.xlsx before applying changes.
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = file_path.with_name(
        f"{file_path.stem}_BACKUP_{timestamp}{file_path.suffix}"
    )
    shutil.copy2(file_path, backup_path)
    return backup_path


def get_or_create_sheet(wb, sheet_name, headers=None):
    """
    Get an existing worksheet or create it if it does not exist.
    Optionally write headers if the sheet is empty.
    """
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)

    if headers:
        has_any_header = any(
            ws.cell(row=1, column=col).value
            for col in range(1, len(headers) + 1)
        )

        if not has_any_header:
            for col, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col).value = header

    return ws


def ensure_base_headers(ws_data, ws_not_found):
    """
    Ensure the four base headers are German and consistent in Source_Data
    and Not_Found.
    """
    base_headers = [
        "Interne ID",
        "Unternehmensname",
        "Ort",
        "Nationale ID",
    ]

    for col, header in enumerate(base_headers, start=1):
        ws_data.cell(row=1, column=col).value = header
        ws_not_found.cell(row=1, column=col).value = header


def append_log(
    ws_log,
    level,
    source_sheet,
    source_row,
    internal_id,
    name,
    national_id,
    database_id,
    reason,
):
    """
    Append one entry to the Apply_Log worksheet.

    One log row is written per processed Search_Export row.
    Previous names, when found, are included in the same log entry.
    """
    next_row = ws_log.max_row + 1

    ws_log.cell(row=next_row, column=1).value = datetime.now().strftime("%Y-%m-%d")
    ws_log.cell(row=next_row, column=2).value = datetime.now().strftime("%H:%M:%S")
    ws_log.cell(row=next_row, column=3).value = level
    ws_log.cell(row=next_row, column=4).value = source_sheet
    ws_log.cell(row=next_row, column=5).value = source_row
    ws_log.cell(row=next_row, column=6).value = clean_id_or_database_id_for_match(internal_id)
    ws_log.cell(row=next_row, column=7).value = clean_outer_quotes(name)
    ws_log.cell(row=next_row, column=8).value = clean_national_id_for_match(national_id)
    ws_log.cell(row=next_row, column=9).value = clean_id_or_database_id_for_match(database_id)
    ws_log.cell(row=next_row, column=10).value = reason


# ============================================================
# NATIONALE ID CLEANING FUNCTION
# ============================================================

def remove_all_spaces_from_national_id_column_d(ws_data):
    """
    Remove all whitespace characters from column D, Nationale ID,
    in Source_Data.

    Examples:
    DE 101 000 001 -> DE101000001
    12 / 67 / 19 -> 12/67/19

    Slashes and other non-whitespace characters are preserved.

    This function runs before building indexes and before any matching.
    """
    changed_count = 0

    for row in range(2, ws_data.max_row + 1):
        cell = ws_data.cell(row=row, column=TARGET_COL_NATIONAL_ID)
        value = cell.value

        if value is None:
            continue

        original = str(value)
        cleaned = re.sub(r"\s+", "", original)

        if cleaned != original:
            cell.value = cleaned
            changed_count += 1

    return changed_count


# ============================================================
# INDEX FUNCTIONS
# ============================================================

def build_target_index(ws_data):
    """
    Index Source_Data by:
    A Interne ID + B Unternehmensname + D Nationale ID.
    """
    index = {}

    for row in range(2, ws_data.max_row + 1):
        key = make_match_key(
            ws_data.cell(row=row, column=TARGET_COL_INTERNAL_ID).value,
            ws_data.cell(row=row, column=TARGET_COL_NAME).value,
            ws_data.cell(row=row, column=TARGET_COL_NATIONAL_ID).value,
        )

        if key not in index:
            index[key] = []

        index[key].append(row)

    return index


def build_batch_index(ws_batch):
    """
    Index Extracted_Batch.xlsx by column D, Datenbank-ID.
    """
    index = {}

    for row in range(2, ws_batch.max_row + 1):
        database_id = clean_id_or_database_id_for_match(
            ws_batch.cell(row=row, column=BATCH_COL_DATABASE_ID).value
        )

        if not database_id:
            continue

        if database_id not in index:
            index[database_id] = []

        index[database_id].append(row)

    return index


# ============================================================
# PREVIOUS NAME FUNCTIONS
# ============================================================

def company_name_title_case(name):
    """
    Convert company names to a more readable title-like casing.

    Example:
    OLD LOGISTICS GMBH
    becomes:
    Old Logistics GmbH
    """
    legal_forms = {
        "GMBH": "GmbH",
        "MBH": "mbH",
        "AG": "AG",
        "KG": "KG",
        "OHG": "OHG",
        "UG": "UG",
        "GBR": "GbR",
        "EG": "eG",
        "EK": "e.K.",
        "E.K.": "e.K.",
        "EV": "e.V.",
        "E.V.": "e.V.",
        "SE": "SE",
        "KGAA": "KGaA",
    }

    parts = re.split(r'(\s+|\+|&|-|/)', name.strip())
    converted = []

    for part in parts:
        upper = part.upper()

        if upper in legal_forms:
            converted.append(legal_forms[upper])
        elif part.strip() == "":
            converted.append(part)
        elif part in ["+", "&", "-", "/"]:
            converted.append(part)
        else:
            converted.append(part[:1].upper() + part[1:].lower())

    return "".join(converted).strip()


def extract_previous_names(value):
    """
    Extract previous company names from Search_Export column H.

    Column H is not used for matching or safety checks.
    It is used only to fill column V, Ehemaliger Name.

    Supported examples:
    - COMPANY GMBH (Previous name: "OLD COMPANY GMBH")
    - COMPANY GMBH (Previous Name: "OLD COMPANY GMBH")
    - COMPANY GMBH (Previous names: "OLD COMPANY GMBH")
    - COMPANY GMBH (Previous name: “OLD COMPANY GMBH”)
    - COMPANY GMBH Previous name: "OLD COMPANY GMBH"
    - COMPANY GMBH (Ehemaliger Name: "OLD COMPANY GMBH")
    - Multiple previous names in the same cell
    """
    if value is None:
        return None

    text = str(value)

    patterns = [
        r'Previous\s+names?\s*:\s*["“”„]([^"“”„]+)["“”„]',
        r'\(Previous\s+names?\s*:\s*["“”„]([^"“”„]+)["“”„]\)',
        r'Previous\s+names?\s*:\s*([^;|,)]+)',
        r'Ehemaliger\s+Name\s*:\s*["“”„]([^"“”„]+)["“”„]',
        r'\(Ehemaliger\s+Name\s*:\s*["“”„]([^"“”„]+)["“”„]\)',
        r'Ehemalige\s+Namen\s*:\s*["“”„]([^"“”„]+)["“”„]',
        r'\(Ehemalige\s+Namen\s*:\s*["“”„]([^"“”„]+)["“”„]\)',
    ]

    found_names = []

    for pattern in patterns:
        matches = re.findall(pattern, text, flags=re.IGNORECASE)

        for match in matches:
            cleaned = collapse_whitespace(match)
            cleaned = cleaned.strip(" .;|,()[]{}")

            if cleaned:
                formatted = company_name_title_case(cleaned)

                if formatted not in found_names:
                    found_names.append(formatted)

    if not found_names:
        return None

    return " | ".join(found_names)


# ============================================================
# DATA COPY AND FORMATTING FUNCTIONS
# ============================================================

def copy_batch_c_to_s_into_target_e_to_u(ws_batch, batch_row, ws_data, target_row):
    """
    Copy Extracted_Batch columns C:S into Source_Data columns E:U.

    Mapping:
    C -> E
    D -> F
    E -> G
    F -> H
    G -> I
    H -> J
    I -> K
    J -> L
    K -> M
    L -> N
    M -> O
    N -> P
    O -> Q
    P -> R
    Q -> S
    R -> T
    S -> U

    This function resets the target cell value and number format
    for the row currently processed.
    """
    source_start_col = 3   # C
    source_end_col = 19    # S
    target_start_col = 5   # E

    for offset, source_col in enumerate(range(source_start_col, source_end_col + 1)):
        source_cell = ws_batch.cell(row=batch_row, column=source_col)
        target_cell = ws_data.cell(row=target_row, column=target_start_col + offset)

        target_cell.value = source_cell.value
        target_cell.number_format = source_cell.number_format


def normalize_turnover_cell(cell):
    """
    Normalize one Umsatz cell in Source_Data column P.

    Rules:
    - n.v. becomes Not found;
    - numeric values are multiplied by 1000;
    - euro formatting is applied;
    - horizontal alignment is set to left.

    This function must only be applied to rows updated in the current run.
    """
    value = cell.value

    cell.alignment = Alignment(horizontal="left")

    if value is None:
        return

    if isinstance(value, str):
        text = value.strip()

        if text.lower() == "n.v.":
            cell.value = "Not found"
            return

        if "€" in text:
            return

        normalized = text.replace(".", "").replace(",", ".")

        try:
            numeric_value = float(normalized)
        except ValueError:
            return

        cell.value = numeric_value * 1000
        cell.number_format = EURO_FORMAT
        return

    if isinstance(value, (int, float)):
        cell.value = value * 1000
        cell.number_format = EURO_FORMAT


def normalize_turnover_only_updated_rows(ws_data, updated_rows):
    """
    Normalize Umsatz only for rows updated in the current run.

    This prevents multiplying old values again when the script is run
    cumulatively in blocks.
    """
    for row in sorted(set(updated_rows)):
        cell = ws_data.cell(row=row, column=16)  # P
        normalize_turnover_cell(cell)


def align_column_p_left(ws_data):
    """
    Set horizontal alignment of column P to left without changing values.
    This may be applied globally because it does not change cell values.
    """
    for row in range(1, ws_data.max_row + 1):
        ws_data.cell(row=row, column=16).alignment = Alignment(horizontal="left")


def ensure_target_headers(ws_data, ws_batch):
    """
    Ensure that columns E:U in Source_Data have the same headers as columns C:S
    in Extracted_Batch.xlsx.

    Also ensure columns V and W have the required German headers.
    """
    target_col = TARGET_PASTE_START_COL

    for source_col in range(3, 20):  # C:S
        header = ws_batch.cell(row=1, column=source_col).value
        ws_data.cell(row=1, column=target_col).value = header
        target_col += 1

    ws_data.cell(row=1, column=TARGET_COL_FORMER_NAME).value = "Ehemaliger Name"
    ws_data.cell(row=1, column=TARGET_COL_DUP_GROUP).value = "Duplikatgruppen-ID"


def move_rows_to_not_found(ws_data, ws_not_found, rows_to_move):
    """
    Move rows from Source_Data to Not_Found.

    Only columns A:D are copied, because Not_Found has the original
    four-column structure:
    Interne ID, Unternehmensname, Ort, Nationale ID.

    Rows are deleted from Source_Data after being copied.
    """
    if not rows_to_move:
        return

    rows_to_move = sorted(set(rows_to_move))

    for row in rows_to_move:
        next_row = ws_not_found.max_row + 1

        for col in range(1, 5):  # A:D
            ws_not_found.cell(row=next_row, column=col).value = ws_data.cell(row=row, column=col).value

    for row in sorted(rows_to_move, reverse=True):
        ws_data.delete_rows(row, 1)


# ============================================================
# DUPLICATE GROUP FUNCTIONS
# ============================================================

def assign_duplicate_group_ids(ws_data):
    """
    Assign Duplikatgruppen-IDs in column W.

    Duplicate detection is cumulative.
    The function scans all rows currently present in Source_Data,
    not only the rows from the current Search_Export block.

    A duplicate group is created only if:
    - column F, Datenbank-ID, is identical;
    - column Q, UID-Nummer, is identical;
    - column F is not empty;
    - column Q is not empty;
    - the group contains at least two different internal IDs in column A.

    Company name, city, Nationale ID and register number are NOT used
    for duplicate detection.

    This function only writes to column W.
    It does not change A:V.
    It does not move rows.
    It does not delete rows.
    """
    for row in range(2, ws_data.max_row + 1):
        ws_data.cell(row=row, column=TARGET_COL_DUP_GROUP).value = None

    groups = {}

    for row in range(2, ws_data.max_row + 1):
        internal_id = clean_id_or_database_id_for_match(
            ws_data.cell(row=row, column=TARGET_COL_INTERNAL_ID).value
        )

        database_id = clean_id_or_database_id_for_match(
            ws_data.cell(row=row, column=6).value  # F = Datenbank-ID
        )

        uid_number = clean_national_id_for_match(
            ws_data.cell(row=row, column=17).value  # Q = UID-Nummer
        )

        if not internal_id or not database_id or not uid_number:
            continue

        key = (database_id, uid_number)

        if key not in groups:
            groups[key] = []

        groups[key].append((row, internal_id))

    duplicate_counter = 1

    for rows_with_ids in groups.values():
        if len(rows_with_ids) < 2:
            continue

        distinct_internal_ids = {
            internal_id
            for _, internal_id in rows_with_ids
            if internal_id
        }

        if len(distinct_internal_ids) < 2:
            continue

        duplicate_id = f"DUP_{duplicate_counter:06d}"

        for row, _ in rows_with_ids:
            ws_data.cell(row=row, column=TARGET_COL_DUP_GROUP).value = duplicate_id

        duplicate_counter += 1

    return duplicate_counter - 1


# ============================================================
# MAIN SCRIPT
# ============================================================

def main():
    if not SAMPLE_DATA_DIR.exists():
        raise FileNotFoundError(
            f"Folder not found: {SAMPLE_DATA_DIR}. "
            "Please create sample_data and place the Excel files inside it."
        )

    if not MATCH_FILE.exists():
        raise FileNotFoundError(f"File not found: {MATCH_FILE}")

    if not EXTRACTED_BATCH_FILE.exists():
        raise FileNotFoundError(f"File not found: {EXTRACTED_BATCH_FILE}")

    search_export_file = find_search_export_file()
    backup_path = create_backup(MATCH_FILE)

    print(f"Backup created: {backup_path.name}")
    print(f"Search export file used: {search_export_file.name}")
    print("Cumulative mode: rows not present in the current Search_Export will not be moved.")
    print("Umsatz mode: only rows updated in the current run will be multiplied by 1000.")
    print("Duplicate mode: column W is recalculated globally and cumulatively.")

    wb_match = load_workbook(MATCH_FILE)
    wb_export = load_workbook(search_export_file, data_only=True)
    wb_batch = load_workbook(EXTRACTED_BATCH_FILE, data_only=True)

    if SHEET_DATA not in wb_match.sheetnames:
        raise RuntimeError(f'The worksheet "{SHEET_DATA}" was not found in {MATCH_FILE.name}.')

    if SHEET_NOT_FOUND not in wb_match.sheetnames:
        raise RuntimeError(f'The worksheet "{SHEET_NOT_FOUND}" was not found in {MATCH_FILE.name}.')

    ws_data = wb_match[SHEET_DATA]
    ws_not_found = wb_match[SHEET_NOT_FOUND]
    ws_export = wb_export[wb_export.sheetnames[0]]
    ws_batch = wb_batch[wb_batch.sheetnames[0]]

    ensure_base_headers(ws_data, ws_not_found)

    national_id_spaces_removed_count = remove_all_spaces_from_national_id_column_d(ws_data)

    ws_log = get_or_create_sheet(
        wb_match,
        SHEET_LOG,
        headers=[
            "Datum",
            "Uhrzeit",
            "Status",
            "Quellblatt",
            "Quellzeile",
            "Interne ID",
            "Unternehmensname",
            "Nationale ID",
            "Datenbank-ID",
            "Grund",
        ],
    )

    ensure_target_headers(ws_data, ws_batch)

    target_index = build_target_index(ws_data)
    batch_index = build_batch_index(ws_batch)

    rows_to_move_to_not_found = set()
    updated_target_rows = []

    processed_count = 0
    updated_count = 0
    previous_name_filled_count = 0
    moved_no_database_id_count = 0
    moved_database_id_not_found_in_batch_count = 0
    moved_uid_mismatch_count = 0
    duplicate_target_count = 0
    duplicate_batch_count = 0
    export_without_target_count = 0
    skipped_already_scheduled_to_move_count = 0

    for export_row in range(2, ws_export.max_row + 1):
        export_internal_id = ws_export.cell(row=export_row, column=EXPORT_COL_INTERNAL_ID).value
        export_name = ws_export.cell(row=export_row, column=EXPORT_COL_NAME).value
        export_national_id = ws_export.cell(row=export_row, column=EXPORT_COL_NATIONAL_ID).value
        export_database_id = ws_export.cell(row=export_row, column=EXPORT_COL_DATABASE_ID).value
        export_matched_name = ws_export.cell(row=export_row, column=EXPORT_COL_MATCHED_NAME).value

        clean_internal_id = clean_id_or_database_id_for_match(export_internal_id)
        clean_name = clean_outer_quotes(export_name)
        clean_national_id = clean_national_id_for_match(export_national_id)
        clean_database_id = clean_id_or_database_id_for_match(export_database_id)

        if not clean_internal_id and not normalize_company_name_for_match(export_name) and not clean_national_id:
            continue

        processed_count += 1

        target_key = make_match_key(export_internal_id, export_name, export_national_id)
        target_rows = target_index.get(target_key, [])

        if len(target_rows) == 0:
            export_without_target_count += 1

            append_log(
                ws_log=ws_log,
                level="WARNING",
                source_sheet=ws_export.title,
                source_row=export_row,
                internal_id=clean_internal_id,
                name=clean_name,
                national_id=clean_national_id,
                database_id=clean_database_id,
                reason=(
                    "The row exists in the Search_Export, but no matching row was found in "
                    "Source_Data using A+B+C from Search_Export against A+B+D from Match_Data. "
                    "Company names were normalized for spaces, case and special characters. "
                    "Nationale ID values were compared after removing all whitespace. "
                    "No row was moved because there is no target row in Match_Data to move."
                ),
            )

            continue

        if len(target_rows) > 1:
            duplicate_target_count += 1

            append_log(
                ws_log=ws_log,
                level="ERROR",
                source_sheet=ws_export.title,
                source_row=export_row,
                internal_id=clean_internal_id,
                name=clean_name,
                national_id=clean_national_id,
                database_id=clean_database_id,
                reason=(
                    f"More than one matching row found in Source_Data: {target_rows}. "
                    "No update was applied and no row was moved to prevent an incorrect match."
                ),
            )

            continue

        target_row = target_rows[0]

        if target_row in rows_to_move_to_not_found:
            skipped_already_scheduled_to_move_count += 1

            append_log(
                ws_log=ws_log,
                level="ERROR",
                source_sheet=ws_export.title,
                source_row=export_row,
                internal_id=clean_internal_id,
                name=clean_name,
                national_id=clean_national_id,
                database_id=clean_database_id,
                reason=(
                    f"The matching target row {target_row} was already scheduled to be moved "
                    "to Not_Found during this run. No further action was applied "
                    "to prevent inconsistent processing."
                ),
            )

            continue

        if not clean_database_id:
            moved_no_database_id_count += 1
            rows_to_move_to_not_found.add(target_row)

            append_log(
                ws_log=ws_log,
                level="INFO",
                source_sheet=ws_export.title,
                source_row=export_row,
                internal_id=clean_internal_id,
                name=clean_name,
                national_id=clean_national_id,
                database_id=clean_database_id,
                reason=(
                    "The row was matched by A+B+C from Search_Export against A+B+D from Match_Data, "
                    "but column G, Gematchte Datenbank-ID, is empty. This means the company was not found "
                    "in the external data source. The row will be moved to Not_Found."
                ),
            )

            continue

        batch_rows = batch_index.get(clean_database_id, [])

        if len(batch_rows) == 0:
            moved_database_id_not_found_in_batch_count += 1
            rows_to_move_to_not_found.add(target_row)

            append_log(
                ws_log=ws_log,
                level="ERROR",
                source_sheet=ws_export.title,
                source_row=export_row,
                internal_id=clean_internal_id,
                name=clean_name,
                national_id=clean_national_id,
                database_id=clean_database_id,
                reason=(
                    "Datenbank-ID check failed: the Datenbank-ID from Search_Export column G was not found "
                    "in Extracted_Batch column D. The row will be moved to Not_Found."
                ),
            )

            continue

        if len(batch_rows) > 1:
            duplicate_batch_count += 1
            rows_to_move_to_not_found.add(target_row)

            append_log(
                ws_log=ws_log,
                level="ERROR",
                source_sheet=ws_export.title,
                source_row=export_row,
                internal_id=clean_internal_id,
                name=clean_name,
                national_id=clean_national_id,
                database_id=clean_database_id,
                reason=(
                    f"Datenbank-ID check failed: the Datenbank-ID from Search_Export column G was found more "
                    f"than once in Extracted_Batch column D: {batch_rows}. "
                    "The row will be moved to Not_Found."
                ),
            )

            continue

        batch_row = batch_rows[0]

        batch_uid = ws_batch.cell(row=batch_row, column=BATCH_COL_UID).value
        target_national_id = ws_data.cell(row=target_row, column=TARGET_COL_NATIONAL_ID).value

        clean_batch_uid = clean_national_id_for_match(batch_uid)
        clean_target_national_id = clean_national_id_for_match(target_national_id)

        if clean_batch_uid != clean_target_national_id:
            moved_uid_mismatch_count += 1
            rows_to_move_to_not_found.add(target_row)

            append_log(
                ws_log=ws_log,
                level="ERROR",
                source_sheet=ws_export.title,
                source_row=export_row,
                internal_id=clean_internal_id,
                name=clean_name,
                national_id=clean_national_id,
                database_id=clean_database_id,
                reason=(
                    "UID safety check failed: Extracted_Batch column O does not match "
                    "Match_Data column D after removing all whitespace from Nationale ID / UID-Nummer values. "
                    f"Extracted_Batch column O = {clean_batch_uid}; "
                    f"Match_Data column D = {clean_target_national_id}. "
                    "The row will be moved to Not_Found."
                ),
            )

            continue

        copy_batch_c_to_s_into_target_e_to_u(
            ws_batch=ws_batch,
            batch_row=batch_row,
            ws_data=ws_data,
            target_row=target_row,
        )

        updated_target_rows.append(target_row)

        previous_names = extract_previous_names(export_matched_name)

        if previous_names:
            ws_data.cell(row=target_row, column=TARGET_COL_FORMER_NAME).value = previous_names
            previous_name_filled_count += 1
            previous_name_log_text = f" Ehemaliger Name written to column V: {previous_names}."
        else:
            previous_name_log_text = " No previous name found in column H."

        updated_count += 1

        append_log(
            ws_log=ws_log,
            level="OK",
            source_sheet=ws_export.title,
            source_row=export_row,
            internal_id=clean_internal_id,
            name=clean_name,
            national_id=clean_national_id,
            database_id=clean_database_id,
            reason=(
                "Row updated successfully. A+B+C from Search_Export matched A+B+D from Match_Data. "
                f"Datenbank-ID from Search_Export column G matched Extracted_Batch row {batch_row}. "
                f"Extracted_Batch column O matched Match_Data column D after removing all whitespace "
                f"from Nationale ID / UID-Nummer values. Data copied from Extracted_Batch C:S to Match_Data E:U "
                f"in row {target_row}. Umsatz normalization will be applied only to this updated row."
                f"{previous_name_log_text}"
            ),
        )

    normalize_turnover_only_updated_rows(ws_data, updated_target_rows)

    move_rows_to_not_found(
        ws_data=ws_data,
        ws_not_found=ws_not_found,
        rows_to_move=rows_to_move_to_not_found,
    )

    align_column_p_left(ws_data)

    duplicate_groups_count = assign_duplicate_group_ids(ws_data)

    wb_match.save(MATCH_FILE)

    print("")
    print("Process completed.")
    print(f"Backup created: {backup_path.name}")
    print(f"Search export file used: {search_export_file.name}")
    print("Cumulative mode: active")
    print("Rows not present in the current Search_Export were not moved.")
    print("Umsatz mode: only rows updated in the current run were multiplied by 1000.")
    print("Duplicate mode: only column W was recalculated globally.")
    print(f"Nationale ID values cleaned in Match_Data column D: {national_id_spaces_removed_count}")
    print(f"Rows processed in Search_Export: {processed_count}")
    print(f"Rows updated in Source_Data: {updated_count}")
    print(f"Rows with Umsatz normalized in this run: {len(set(updated_target_rows))}")
    print(f"Previous names written to column V: {previous_name_filled_count}")
    print(f"Rows moved to Not_Found because column G was empty: {moved_no_database_id_count}")
    print(f"Rows moved to Not_Found because Datenbank-ID was not found in Extracted_Batch: {moved_database_id_not_found_in_batch_count}")
    print(f"Rows moved to Not_Found because UID-Nummer did not match Nationale ID: {moved_uid_mismatch_count}")
    print(f"Rows moved to Not_Found in total: {len(rows_to_move_to_not_found)}")
    print(f"Cases with multiple matching rows in Match_Data: {duplicate_target_count}")
    print(f"Cases with duplicated Datenbank-ID in Extracted_Batch: {duplicate_batch_count}")
    print(f"Rows skipped because target row was already scheduled to move: {skipped_already_scheduled_to_move_count}")
    print(f"Rows present in Search_Export but not found in Match_Data: {export_without_target_count}")
    print(f"Duplicate groups created in column W: {duplicate_groups_count}")
    print("Duplicate detection rule: same Datenbank-ID in column F + same UID-Nummer in column Q + different internal IDs in column A.")
    print("Column P alignment set to left.")
    print(f"File saved: {MATCH_FILE}")


if __name__ == "__main__":
    main()